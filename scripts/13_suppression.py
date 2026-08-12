#!/usr/bin/env python3
"""Map which source cells are privacy-suppressed (blank in the export) rather
than true zeros. The InEK DatenBrowser aggregates only cells with five or more
cases; cells with 1-4 cases are suppressed (returned blank). The extraction
step reads blanks as 0, so downstream tables must distinguish a suppressed cell
(displayed as "<5") from a genuine zero.

Re-reads the raw workbook and writes suppressed_cells.json: for each disposition
category, comorbidity category, and procedure, a per-age-band boolean flag that
is True where the source cell is blank (suppressed). In this export every zero
in the analysed cells is in fact a suppressed blank; no genuine zeros occur.
"""
import openpyxl, json

import glob as _glob, sys as _sys
_wbs = sorted(_glob.glob("data/*.xlsx"))
if not _wbs:
    _sys.exit("No export found. Place your InEK DatenBrowser C71 export (.xlsx) in data/ (see data/README.md).")
WB = _wbs[0]
BANDS = ["18-29", "30-39", "40-49", "50-54", "55-59", "60-64", "65-74", "75-79", ">=80"]
UNION_COLS = [3, 5, 7, 9, 11, 13, 15, 17, 19]

wb = openpyxl.load_workbook(WB, data_only=True)
ov = wb["Overview"]; ch = wb["Charlson Comordity Index"]; up = wb["Übersicht-Prozeduren"]

# Overview main-diagnosis blocks (case row) and disposition columns (from 01)
BLOCKS = {">=80": 7, "75-79": 17, "65-74": 28, "60-64": 39, "55-59": 50,
          "50-54": 61, "40-49": 72, "30-39": 83, "18-29": 94}
DISP_COL = {"death": 11, "hospice": 20, "home": 29, "nursing_home": 38, "transfer_hospital": 47}

def blank(v):
    return v is None

disposition = {}
for key, col in DISP_COL.items():
    disposition[key] = {b: blank(ov.cell(row=BLOCKS[b], column=col).value) for b in BANDS}

# Comorbidity category total rows (from 01)
CHARLSON = [("Other solid tumour (non-C71)", 8), ("Malignant haematological disease", 28),
            ("Diabetes mellitus", 35), ("Dementia", 51), ("Heart failure", 61),
            ("Cerebrovascular disease", 76), ("Chronic pulmonary disease", 101),
            ("Peripheral arterial occlusive disease", 110), ("Collagenoses", 121),
            ("Renal failure / chronic kidney disease", 127)]
comorbidity = {}
for name, row in CHARLSON:
    comorbidity[name] = {b: blank(ch.cell(row=row, column=UNION_COLS[i]).value)
                         for i, b in enumerate(BANDS)}

# Procedures: single codes (grouped families are sums, treated as non-suppressed
# unless every summed sub-code is blank; here we flag the exact single codes).
def proc_row(code):
    for row in range(1, up.max_row + 1):
        if str(up.cell(row=row, column=1).value).strip() == code:
            return row
    return None

PROC_SINGLE = {"Native cranial CT (3-200)": "3-200", "Native cranial MRI (3-800)": "3-800",
               "Microsurgical technique (5-984)": "5-984",
               "Fluorescence-guided surgery (5-989)": "5-989",
               "External ventricular drain (5-022.00)": "5-022.00"}
procedures = {}
for name, code in PROC_SINGLE.items():
    rr = proc_row(code)
    procedures[name] = {b: (blank(up.cell(row=rr, column=UNION_COLS[i]).value) if rr else False)
                        for i, b in enumerate(BANDS)}

out = {"bands": BANDS, "disposition": disposition,
       "comorbidity": comorbidity, "procedures": procedures}
json.dump(out, open("derived_data/suppressed_cells.json", "w"), indent=2)

nsupp = (sum(v for d in disposition.values() for v in d.values())
         + sum(v for d in comorbidity.values() for v in d.values())
         + sum(v for d in procedures.values() for v in d.values()))
print(f"suppressed cells flagged: {nsupp}")
for name, d in comorbidity.items():
    s = [b for b in BANDS if d[b]]
    if s:
        print(f"  comorbidity '{name}': suppressed in {s}")
for key, d in disposition.items():
    s = [b for b in BANDS if d[b]]
    if s:
        print(f"  disposition '{key}': suppressed in {s}")
