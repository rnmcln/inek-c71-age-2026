#!/usr/bin/env python3
"""
Marginal-based mean Charlson comorbidity score by age band.

The workbook provides marginal counts of episodes carrying each comorbidity
category by age band, not the joint distribution within episodes. Because the
mean of a sum equals the sum of the means irrespective of co-occurrence, the
MEAN comorbidity-weighted Charlson score per episode is computable exactly from
these marginals; a patient-level CCI distribution (categories, median) is not.

Weights follow Charlson et al. 1987 (CHV_charlson_index_definition):
  1: myocardial infarction, heart failure, peripheral vascular disease,
     cerebrovascular disease, dementia, chronic pulmonary disease, connective
     tissue disease, ulcer disease, mild liver disease, uncomplicated diabetes
  2: complicated diabetes (end-organ damage), moderate/severe renal disease,
     second non-metastatic solid tumour, leukaemia/lymphoma, hemiplegia
  3: moderate/severe liver disease
  6: metastatic solid tumour, AIDS

Diabetes is split into uncomplicated (weight 1) and complicated (weight 2) by
ICD-10-GM subcode; the de-duplicated category total is apportioned by the
code-level share so episodes are not double-counted. The second-solid-tumour
category is scored directly at code level: extracranial secondary neoplasms
(C77, C78, C80, and C79 other than CNS sites) score as metastatic solid tumour
(weight 6) and other second solid tumours score weight 2. Secondary neoplasms of
the brain and other nervous system (C79.3, C79.4) are excluded from the tumour
comorbidity, because in a primary brain-tumour (C71) cohort these most plausibly
reflect CNS involvement of the index tumour rather than a separate metastatic
cancer. Age points are not added, to avoid circularity in an age-stratified
analysis.
"""
import openpyxl, csv, json, math

import glob as _glob, sys as _sys
_wbs = sorted(_glob.glob("data/*.xlsx"))
if not _wbs:
    _sys.exit("No export found. Place your InEK DatenBrowser C71 export (.xlsx) in data/ (see data/README.md).")
WB = _wbs[0]
BANDS = ["18-29", "30-39", "40-49", "50-54", "55-59", "60-64", "65-74", "75-79", ">=80"]
COLS = [3, 5, 7, 9, 11, 13, 15, 17, 19]
SCORE = [23.5, 34.5, 44.5, 52.0, 57.0, 62.0, 69.5, 77.0, 85.0]

wb = openpyxl.load_workbook(WB, data_only=True)
ch = wb["Charlson Comordity Index"]


def r(x):
    return int(round(float(x))) if x is not None else 0


# category header row -> (name, fixed weight or None if special)
CATS = {
    5: ("HIV/AIDS", 6),
    7: ("Second solid tumour", "TUMOUR"),
    27: ("Leukaemia/lymphoma", 2),
    34: ("Diabetes", "DIABETES"),
    50: ("Dementia", 1),
    60: ("Heart failure", 1),
    72: ("Myocardial infarction", 1),
    75: ("Cerebrovascular disease", 1),
    100: ("Chronic pulmonary disease", 1),
    109: ("Peripheral vascular disease", 1),
    115: ("Ulcer disease", 1),
    117: ("Mild liver disease", 1),
    120: ("Connective tissue disease", 1),
    126: ("Renal disease", 2),
}
TOTAL_ROW = {5: None, 7: 8, 27: 28, 34: 35, 50: 51, 60: 61, 72: None,
             75: 76, 100: 101, 109: 110, 115: None, 117: None, 120: 121, 126: 127}
header_rows = sorted(CATS)


def code_rows(hrow):
    """Rows carrying an ICD code between this header and the next."""
    idx = header_rows.index(hrow)
    nxt = header_rows[idx + 1] if idx + 1 < len(header_rows) else ch.max_row + 1
    out = []
    for row in range(hrow + 1, nxt):
        code = ch.cell(row=row, column=1).value
        b = ch.cell(row=row, column=2).value
        if code and str(code).strip() not in ("total",) and b != "total":
            out.append((str(code).strip(), row))
    return out


def band_counts(row):
    return [r(ch.cell(row=row, column=c).value) for c in COLS]


N = band_counts(3)  # denominators (union population)

points = [0.0] * 9
for hrow, (name, wt) in CATS.items():
    crows = code_rows(hrow)
    trow = TOTAL_ROW[hrow]
    total = band_counts(trow) if trow else None
    if wt == "DIABETES":
        # split complicated vs uncomplicated by subcode (Exx.9x = uncomplicated)
        comp = [0]*9; unc = [0]*9
        for code, row in crows:
            cc = band_counts(row)
            after = code.split(".")[1] if "." in code else "9"
            weightclass = unc if after.startswith("9") else comp
            for i in range(9):
                weightclass[i] += cc[i]
        allc = [comp[i] + unc[i] for i in range(9)]
        for i in range(9):
            tot = total[i] if total else allc[i]
            share_comp = comp[i] / allc[i] if allc[i] else 0.0
            points[i] += tot * (share_comp * 2 + (1 - share_comp) * 1)
    elif wt == "TUMOUR":
        # Direct code-level weighted sum. Extracranial secondary neoplasms
        # (C77, C78, C80, and C79 except CNS sites) score as metastatic solid
        # tumour (weight 6); other second solid tumours score weight 2.
        # Secondary neoplasms of the brain / other nervous system (C79.3, C79.4)
        # are NOT counted: in a primary brain-tumour cohort these most plausibly
        # reflect CNS involvement of the index tumour, not a separate metastatic
        # cancer.
        EXCLUDE = ("C79.3", "C79.4")
        for code, row in crows:
            if any(code.startswith(x) for x in EXCLUDE):
                continue
            cc = band_counts(row)
            metastatic = code[:3] in ("C77", "C78", "C79", "C80")
            weight = 6 if metastatic else 2
            for i in range(9):
                points[i] += weight * cc[i]
    else:
        if total:
            vals = total
        else:  # no total row: sum code rows
            vals = [0]*9
            for code, row in crows:
                cc = band_counts(row)
                for i in range(9):
                    vals[i] += cc[i]
        for i in range(9):
            points[i] += wt * vals[i]

mean_cci = [points[i] / N[i] for i in range(9)]

# Poisson log-linear trend test on rounded comorbidity points, exposure = N,
# covariate = age midpoint. Newton-Raphson; Wald test on slope.
y = [round(p) for p in points]
x = SCORE
logN = [math.log(n) for n in N]
b0, b1 = math.log(sum(y) / sum(N)), 0.0
for _ in range(50):
    g0 = g1 = h00 = h01 = h11 = 0.0
    for i in range(9):
        mu = math.exp(b0 + b1 * x[i] + logN[i])
        g0 += y[i] - mu
        g1 += x[i] * (y[i] - mu)
        h00 -= mu
        h01 -= x[i] * mu
        h11 -= x[i] * x[i] * mu
    det = h00 * h11 - h01 * h01
    db0 = (h11 * g0 - h01 * g1) / det
    db1 = (-h01 * g0 + h00 * g1) / det
    b0 -= db0
    b1 -= db1
    if abs(db0) < 1e-10 and abs(db1) < 1e-10:
        break
# variance of b1 from inverse negative Hessian (Fisher info)
info00 = info01 = info11 = 0.0
for i in range(9):
    mu = math.exp(b0 + b1 * x[i] + logN[i])
    info00 += mu; info01 += x[i] * mu; info11 += x[i] * x[i] * mu
detI = info00 * info11 - info01 * info01
var_b1 = info00 / detI
se_b1 = math.sqrt(var_b1)
z = b1 / se_b1
p = 2 * (1 - 0.5 * (1 + math.erf(abs(z) / math.sqrt(2))))
rr_per_decade = math.exp(b1 * 10)

with open("derived_data/charlson_index_by_age.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["age_band", "n_union", "comorbidity_points", "mean_cci"])
    for i in range(9):
        w.writerow([BANDS[i], N[i], round(points[i], 1), round(mean_cci[i], 3)])

json.dump({"mean_cci": {BANDS[i]: round(mean_cci[i], 3) for i in range(9)},
           "n_union": {BANDS[i]: N[i] for i in range(9)},
           "poisson_slope_per_year": b1, "rr_per_decade": rr_per_decade,
           "z": z, "p_trend": p},
          open("derived_data/charlson_index.json", "w"), indent=2)

print("Age band   N    mean CCI")
for i in range(9):
    print(f"  {BANDS[i]:>6} {N[i]:>5}   {mean_cci[i]:.2f}")
print(f"\nPoisson trend: rate ratio per decade = {rr_per_decade:.2f}, "
      f"z = {z:.2f}, p = {'<0.001' if p < 1e-3 else f'{p:.3f}'}")
print(f"Overall mean CCI (all ages) = {sum(points)/sum(N):.2f}")
