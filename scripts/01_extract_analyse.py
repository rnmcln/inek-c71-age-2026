#!/usr/bin/env python3
"""
C71 age-stratified analysis: extraction + statistics.

Reads the InEK DatenBrowser age export workbook and produces tidy CSVs and a
statistics JSON. All percentages are recomputed from raw counts; stored
percentages in the workbook are ignored.

Cohorts
-------
Principal-diagnosis cohort (N sums to 18,621): C71 recorded as the principal
(main) hospital discharge diagnosis. Used for age distribution, sex, and
discharge disposition (workbook "main diagnosis" columns).

C71-coded population (N sums to 23,594): C71 recorded as principal OR secondary
diagnosis. This is the denominator to which the comorbidity, secondary-diagnosis
and procedure worksheets are keyed. Used only for those analyses.

Author: analysis pipeline for Lawson McLean et al., C71 age study.
"""
import json
import math
import openpyxl
import numpy as np

# ---------------------------------------------------------------------------
# Pure-Python statistics (no scipy/statsmodels dependency)
# ---------------------------------------------------------------------------
def norm_cdf(x):
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def _gammln(xx):
    cof = [76.18009172947146, -86.50532032941677, 24.01409824083091,
           -1.231739572450155, 0.1208650973866179e-2, -0.5395239384953e-5]
    x = xx
    y = xx
    tmp = x + 5.5
    tmp -= (x + 0.5) * math.log(tmp)
    ser = 1.000000000190015
    for c in cof:
        y += 1
        ser += c / y
    return -tmp + math.log(2.5066282746310005 * ser / x)


def _gammq(a, x):
    """Regularised upper incomplete gamma Q(a,x) (Numerical Recipes)."""
    if x < 0 or a <= 0:
        raise ValueError
    if x == 0:
        return 1.0
    if x < a + 1.0:  # series for P, then Q = 1 - P
        ap = a
        s = 1.0 / a
        d = s
        for _ in range(1000):
            ap += 1
            d *= x / ap
            s += d
            if abs(d) < abs(s) * 1e-12:
                break
        return 1.0 - s * math.exp(-x + a * math.log(x) - _gammln(a))
    # continued fraction for Q
    b = x + 1.0 - a
    c = 1e300
    d = 1.0 / b
    h = d
    for i in range(1, 1000):
        an = -i * (i - a)
        b += 2.0
        d = an * d + b
        if abs(d) < 1e-300:
            d = 1e-300
        c = b + an / c
        if abs(c) < 1e-300:
            c = 1e-300
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < 1e-12:
            break
    return math.exp(-x + a * math.log(x) - _gammln(a)) * h


def chi2_sf(x, k):
    """Upper tail (survival) of chi-square with k dof."""
    if x <= 0:
        return 1.0
    return _gammq(k / 2.0, x / 2.0)


def chi2_contingency(table):
    """Pearson chi-square test of independence. table: 2D array-like."""
    t = np.asarray(table, float)
    rt = t.sum(axis=1, keepdims=True)
    ct = t.sum(axis=0, keepdims=True)
    tot = t.sum()
    exp = rt @ ct / tot
    chi2 = float(((t - exp) ** 2 / exp).sum())
    dof = (t.shape[0] - 1) * (t.shape[1] - 1)
    return chi2, chi2_sf(chi2, dof), dof


def wilson_ci(k, n, z=1.959963984540054):
    if n == 0:
        return (float("nan"), float("nan"))
    p = k / n
    denom = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / denom
    half = (z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))) / denom
    return (centre - half, centre + half)


def bh_fdr(pvals):
    """Benjamini-Hochberg FDR-adjusted p-values."""
    p = np.asarray(pvals, float)
    m = len(p)
    order = np.argsort(p)
    ranked = p[order] * m / (np.arange(m) + 1)
    # enforce monotonicity
    ranked = np.minimum.accumulate(ranked[::-1])[::-1]
    adj = np.empty(m)
    adj[order] = np.clip(ranked, 0, 1)
    return adj

import glob as _glob, sys as _sys
_wbs = sorted(_glob.glob("data/*.xlsx"))
if not _wbs:
    _sys.exit("No export found. Place your InEK DatenBrowser C71 export (.xlsx) in data/ (see data/README.md).")
WB = _wbs[0]
OUT = "derived_data"

# Ordered age bands and continuous scores (midpoints; >80 -> 85) for trend tests
AGE_BANDS = ["18-29", "30-39", "40-49", "50-54", "55-59", "60-64",
             "65-74", "75-79", ">=80"]
AGE_SCORE = {"18-29": 23.5, "30-39": 34.5, "40-49": 44.5, "50-54": 52.0,
             "55-59": 57.0, "60-64": 62.0, "65-74": 69.5, "75-79": 77.0,
             ">=80": 85.0}

wb = openpyxl.load_workbook(WB, data_only=True)


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return v


def r(x):
    """Round workbook floats (some carry FP noise / privacy rounding) to int."""
    if x is None:
        return 0
    return int(round(float(x)))


# ---------------------------------------------------------------------------
# 1. Principal-diagnosis cohort: Overview sheet
# ---------------------------------------------------------------------------
ov = wb["Overview"]
# (cases_row, male_row, female_row) per band, in workbook order (>=80 first)
BLOCKS = [(">=80", 7, 10, 11), ("75-79", 17, 20, 21), ("65-74", 28, 31, 32),
          ("60-64", 39, 42, 43), ("55-59", 50, 53, 54), ("50-54", 61, 64, 65),
          ("40-49", 72, 75, 76), ("30-39", 83, 86, 87), ("18-29", 94, 97, 98)]
# columns (1-indexed): main-diagnosis column for each disposition
COL = {"all_main": 2, "all_total": 8,
       "death_main": 11, "death_total": 17,
       "hospice_main": 20, "hospice_total": 26,
       "home_main": 29, "home_total": 35,
       "nursing_main": 38, "nursing_total": 44,
       "transfer_main": 47, "transfer_total": 53,
       "other_total": 56}

principal = {}
for band, cr, mr, fr in BLOCKS:
    n = r(cell(ov, cr, COL["all_main"]))
    male = r(cell(ov, mr, COL["all_main"]))
    female = r(cell(ov, fr, COL["all_main"]))
    death = r(cell(ov, cr, COL["death_main"]))
    hospice = r(cell(ov, cr, COL["hospice_main"]))
    home = r(cell(ov, cr, COL["home_main"]))
    nursing = r(cell(ov, cr, COL["nursing_main"]))
    transfer = r(cell(ov, cr, COL["transfer_main"]))
    other = n - (death + hospice + home + nursing + transfer)
    principal[band] = dict(n=n, male=male, female=female, death=death,
                           hospice=hospice, home=home, nursing=nursing,
                           transfer=transfer, other=max(other, 0))

# order ascending
principal = {b: principal[b] for b in AGE_BANDS}
N_principal = sum(v["n"] for v in principal.values())

# ---------------------------------------------------------------------------
# 2. C71-coded population totals (union) from Charlson row 3 (== procedure row2)
# ---------------------------------------------------------------------------
ch = wb["Charlson Comordity Index"]
# columns for age bands on Charlson/Procedure sheets: C,E,G,I,K,M,O,Q,S
UNION_COLS = [3, 5, 7, 9, 11, 13, 15, 17, 19]  # 18-29 ... >=80
union_tot = {AGE_BANDS[i]: r(cell(ch, 3, UNION_COLS[i])) for i in range(9)}
N_union = sum(union_tot.values())

# ---------------------------------------------------------------------------
# 3. Comorbidities (Charlson), keyed to union population
# ---------------------------------------------------------------------------
CHARLSON = [("Other solid tumour (non-C71)", 8),
            ("Malignant haematological disease", 28),
            ("Diabetes mellitus", 35),
            ("Dementia", 51),
            ("Heart failure", 61),
            ("Cerebrovascular disease", 76),
            ("Chronic pulmonary disease", 101),
            ("Peripheral arterial occlusive disease", 110),
            ("Collagenoses", 121),
            ("Renal failure / chronic kidney disease", 127)]
comorbid = {}
for name, row in CHARLSON:
    comorbid[name] = {AGE_BANDS[i]: r(cell(ch, row, UNION_COLS[i]))
                      for i in range(9)}

# ---------------------------------------------------------------------------
# 4. Procedures (single OPS codes), keyed to union population
#    From the 'Übersicht-Prozeduren' sheet.
# ---------------------------------------------------------------------------
up = wb["Übersicht-Prozeduren"]


def proc_row_by_code(code, occurrence=1):
    seen = 0
    for row in range(1, up.max_row + 1):
        if str(cell(up, row, 1)).strip() == code:
            seen += 1
            if seen == occurrence:
                return row
    return None


def proc_counts(rows):
    """Sum one or more rows (list) across age bands."""
    out = {}
    for i in range(9):
        s = 0
        for row in rows:
            s += r(cell(up, row, UNION_COLS[i]))
        out[AGE_BANDS[i]] = s
    return out


# single, clean OPS codes (each episode counted once per code)
PROC_SINGLE = {
    "Native cranial CT (3-200)": ["3-200"],
    "Native cranial MRI (3-800)": ["3-800"],
    "Microsurgical technique (5-984)": ["5-984"],
    "Fluorescence-guided surgery (5-989)": ["5-989"],
    "External ventricular drain (5-022.00)": ["5-022.00"],
}
procedures = {}
for label, codes in PROC_SINGLE.items():
    rows = []
    for cd in codes:
        rr = proc_row_by_code(cd)
        if rr:
            rows.append(rr)
    procedures[label] = proc_counts(rows)

# summed (approximate) categories, reported with explicit caveat
PROC_SUM = {
    "Any intraoperative navigation (5-988.x)":
        ["5-988.0", "5-988.1", "5-988.2", "5-988.3", "5-988.x"],
    "Stereotactic biopsy, supratentorial (1-511.0x)":
        ["1-511.00", "1-511.01"],
    "Inpatient high-voltage radiotherapy (8-522.x)":
        ["8-522.90", "8-522.91", "8-522.b0", "8-522.b1", "8-522.d0", "8-522.d1"],
}
procedures_sum = {}
for label, codes in PROC_SUM.items():
    rows = []
    for cd in codes:
        rr = proc_row_by_code(cd)
        if rr:
            rows.append(rr)
    procedures_sum[label] = proc_counts(rows)

# ---------------------------------------------------------------------------
# Statistics helpers
# ---------------------------------------------------------------------------
def wilson(k, n):
    return wilson_ci(k, n)


def chi2_across(events, totals):
    """Global chi-square of independence for events vs non-events across bands."""
    tbl = [[e, t - e] for e, t in zip(events, totals)]
    chi2, p, dof = chi2_contingency(tbl)
    return chi2, p, dof


def cochran_armitage(events, totals, scores):
    """Two-sided Cochran-Armitage trend test with continuous scores."""
    ri = np.array(events, float)
    ni = np.array(totals, float)
    xi = np.array(scores, float)
    N = ni.sum()
    R = ri.sum()
    p = R / N
    T = np.sum(xi * (ri - ni * p))
    var = p * (1 - p) * (np.sum(ni * xi ** 2) - (np.sum(ni * xi)) ** 2 / N)
    if var <= 0:
        return float("nan"), float("nan")
    z = T / np.sqrt(var)
    pval = 2 * (1 - norm_cdf(abs(z)))
    return z, pval


def analyse_binary(events_by_band, totals_by_band, label):
    events = [events_by_band[b] for b in AGE_BANDS]
    totals = [totals_by_band[b] for b in AGE_BANDS]
    scores = [AGE_SCORE[b] for b in AGE_BANDS]
    props = [(e / t if t else float("nan")) for e, t in zip(events, totals)]
    cis = [wilson(e, t) for e, t in zip(events, totals)]
    chi2, p_chi, dof = chi2_across(events, totals)
    z, p_trend = cochran_armitage(events, totals, scores)
    return dict(label=label, bands=AGE_BANDS, events=events, totals=totals,
                prop=props, ci_low=[c[0] for c in cis],
                ci_high=[c[1] for c in cis],
                chi2=chi2, dof=dof, p_chi2=p_chi, z_trend=z, p_trend=p_trend)


results = {"meta": {"N_principal": N_principal, "N_union": N_union,
                    "principal_by_band": {b: principal[b]["n"] for b in AGE_BANDS},
                    "union_by_band": union_tot}}

# --- sex (male proportion), principal cohort
results["sex_male"] = analyse_binary(
    {b: principal[b]["male"] for b in AGE_BANDS},
    {b: principal[b]["n"] for b in AGE_BANDS}, "Male proportion")

# --- disposition, principal cohort
tot_principal = {b: principal[b]["n"] for b in AGE_BANDS}
for key, lab in [("death", "In-hospital mortality"),
                 ("home", "Discharged home"),
                 ("nursing", "Transfer to nursing home"),
                 ("hospice", "Transfer to hospice care"),
                 ("transfer", "Transfer to another hospital")]:
    results[f"disp_{key}"] = analyse_binary(
        {b: principal[b][key] for b in AGE_BANDS}, tot_principal, lab)

# --- comorbidity, union population
comorbid_stats = {}
p_trends = []
keys = []
for name, byband in comorbid.items():
    st = analyse_binary(byband, union_tot, name)
    comorbid_stats[name] = st
    p_trends.append(st["p_trend"])
    keys.append(name)
# FDR across comorbidity trend tests
p_adj = bh_fdr(p_trends)
for k, pa in zip(keys, p_adj):
    comorbid_stats[k]["p_trend_fdr"] = pa
results["comorbidity"] = comorbid_stats

# --- procedures, union population
proc_stats = {}
p_trends = []
keys = []
allproc = {**procedures, **procedures_sum}
for name, byband in allproc.items():
    st = analyse_binary(byband, union_tot, name)
    proc_stats[name] = st
    p_trends.append(st["p_trend"])
    keys.append(name)
p_adj = bh_fdr(p_trends)
for k, pa in zip(keys, p_adj):
    proc_stats[k]["p_trend_fdr"] = pa
results["procedures"] = proc_stats

# --- overall sex-by-age contingency (global association)
male = [principal[b]["male"] for b in AGE_BANDS]
female = [principal[b]["female"] for b in AGE_BANDS]
chi2, p, dof = chi2_contingency([male, female])
results["sex_global"] = dict(chi2=chi2, dof=dof, p=p)

# --- overall disposition-by-age contingency (6 x 9)
disp_matrix = []
for key in ["death", "hospice", "home", "nursing", "transfer", "other"]:
    disp_matrix.append([principal[b][key] for b in AGE_BANDS])
chi2, p, dof = chi2_contingency(disp_matrix)
results["disposition_global"] = dict(chi2=chi2, dof=dof, p=p)

# ---------------------------------------------------------------------------
# Write derived CSVs
# ---------------------------------------------------------------------------
import csv

with open(f"{OUT}/principal_cohort_by_age.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["age_band", "n_principal", "male", "female", "death", "hospice",
                "home", "nursing_home", "transfer_hospital", "other_residual"])
    for b in AGE_BANDS:
        d = principal[b]
        w.writerow([b, d["n"], d["male"], d["female"], d["death"], d["hospice"],
                    d["home"], d["nursing"], d["transfer"], d["other"]])

with open(f"{OUT}/union_population_by_age.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["age_band", "n_union"])
    for b in AGE_BANDS:
        w.writerow([b, union_tot[b]])

with open(f"{OUT}/comorbidity_by_age.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["comorbidity"] + AGE_BANDS + ["p_trend", "p_trend_fdr"])
    for name in comorbid:
        st = comorbid_stats[name]
        w.writerow([name] + st["events"] + [st["p_trend"], st["p_trend_fdr"]])

with open(f"{OUT}/procedures_by_age.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["procedure"] + AGE_BANDS + ["p_trend", "p_trend_fdr"])
    for name in allproc:
        st = proc_stats[name]
        w.writerow([name] + st["events"] + [st["p_trend"], st["p_trend_fdr"]])

with open(f"{OUT}/statistics.json", "w") as f:
    json.dump(results, f, indent=2, default=lambda o: None if isinstance(o, float) and np.isnan(o) else o)

# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------
print(f"N principal (C71 main dx) = {N_principal}")
print(f"N union (C71 main or sec) = {N_union}")
print("\nAge distribution (principal):")
for b in AGE_BANDS:
    print(f"  {b:>6}: {principal[b]['n']:>6}  ({100*principal[b]['n']/N_principal:4.1f}%)  "
          f"male {100*principal[b]['male']/principal[b]['n']:4.1f}%")
print(f"\nSex global chi2 p = {results['sex_global']['p']:.3g}; "
      f"male-trend p = {results['sex_male']['p_trend']:.3g}")
print("\nDisposition (principal) — rate by age, chi2 p, trend p:")
for key in ["death", "home", "nursing", "hospice", "transfer"]:
    st = results[f"disp_{key}"]
    rates = "  ".join(f"{100*p:4.1f}" for p in st["prop"])
    print(f"  {st['label']:<30} {rates}   chi2p={st['p_chi2']:.2g} trendp={st['p_trend']:.2g}")
print("\nComorbidity prevalence (union) — youngest vs oldest, trend p(FDR):")
for name, st in comorbid_stats.items():
    print(f"  {name:<42} {100*st['prop'][0]:4.1f}% -> {100*st['prop'][-1]:4.1f}%  "
          f"pFDR={st['p_trend_fdr']:.2g}")
print("\nProcedures (union) — youngest vs oldest, trend p(FDR):")
for name, st in proc_stats.items():
    print(f"  {name:<48} {100*st['prop'][0]:4.1f}% -> {100*st['prop'][-1]:4.1f}%  "
          f"pFDR={st['p_trend_fdr']:.2g}")
